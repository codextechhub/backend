"""File writers - the only place that knows what a CSV or an XLSX looks like.

Kept apart from the engine so a new format is one function here plus one entry in
:data:`~vs_exports.constants.FORMAT_MEDIA`, and so the engine can stay about rows and
permissions. Both writers take an iterable of already-rendered string rows, because
value formatting is a catalogue concern (``people`` vs ``system``) and must not be
re-decided per format.

Options are read from the discriminated ``format_options`` object; unknown keys are
ignored and missing ones fall back to the catalogue's declared default, so a file
written by an older definition still opens.
"""
from __future__ import annotations

import csv
import io

from .catalogue import default_format_options
from .constants import ExportFormat


# Merge caller options over the schema defaults for one format.
def _options(fmt: str, options: dict | None) -> dict:
    merged = default_format_options(fmt)
    merged.update({k: v for k, v in (options or {}).items() if k in merged})
    return merged


# Write rows as CSV bytes.
def write_csv(headers, rows, *, options=None, filters_summary=None) -> bytes:
    """Encode ``rows`` as CSV honouring delimiter, encoding, header and quoting.

    ``filters_summary`` is accepted and ignored: CSV has no second sheet to put it on,
    and silently prepending commentary rows would break the importers this format
    exists for.
    """
    opts = _options(ExportFormat.CSV, options)
    buf = io.StringIO(newline="")
    writer = csv.writer(
        buf,
        delimiter=opts["delimiter"],
        quoting=csv.QUOTE_ALL if opts["quote_all"] else csv.QUOTE_MINIMAL,
        lineterminator=opts["line_ending"],
    )
    if opts["header_row"]:
        writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    # errors="replace" so one unmappable character in a legacy encoding cannot fail a
    # run that has already read every row.
    return buf.getvalue().encode(opts["encoding"], errors="replace")


# Write rows as an Excel workbook.
def write_xlsx(headers, rows, *, options=None, filters_summary=None) -> bytes:
    """Encode ``rows`` as a single-sheet workbook, optionally with a filters sheet.

    The filters sheet is recommended whenever a file leaves the finance team: it is
    what lets a reader six weeks later know which slice of the data they are holding.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    opts = _options(ExportFormat.XLSX, options)
    wb = Workbook(write_only=False)
    ws = wb.active
    # Excel rejects sheet names over 31 chars and a handful of punctuation marks.
    ws.title = "".join(c for c in str(opts["sheet_name"]) if c not in "[]:*?/\\")[:31] or "Export"

    bold = Font(bold=True)
    ws.append(list(headers))
    for cell in ws[1]:
        cell.font = bold

    widest = [len(str(h)) for h in headers]
    for row in rows:
        ws.append(list(row))
        if opts["auto_width"]:
            for i, value in enumerate(row):
                if i < len(widest):
                    widest[i] = max(widest[i], len(str(value)))

    if opts["freeze_header"]:
        ws.freeze_panes = "A2"

    if opts["auto_width"]:
        for i, width in enumerate(widest, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 48)

    if opts["filters_sheet"] and filters_summary:
        meta = wb.create_sheet("Filters")
        meta.append(["What produced this file"])
        meta["A1"].font = bold
        for label, value in filters_summary:
            meta.append([label, value])
        meta.column_dimensions["A"].width = 28
        meta.column_dimensions["B"].width = 72

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


_WRITERS = {ExportFormat.CSV: write_csv, ExportFormat.XLSX: write_xlsx}


# Render rows in the requested format.
def write(fmt: str, headers, rows, *, options=None, filters_summary=None) -> bytes:
    """Dispatch to the writer for ``fmt``.

    Raises :class:`ValueError` for a format this version cannot produce (PDF and JSON
    are explicitly post-v1), which the caller turns into a validation error.
    """
    writer = _WRITERS.get(fmt)
    if writer is None:
        raise ValueError(
            f"'{fmt}' is not a format this version can produce. Choose "
            f"{' or '.join(_WRITERS)}."
        )
    return writer(headers, rows, options=options, filters_summary=filters_summary)
