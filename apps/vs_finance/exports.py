"""Tabular exports for finance reports — CSV, Excel (xlsx) and PDF.

A report is reduced to a neutral :class:`ReportTable` (title + column headers + rows,
with optional bold *summary* rows for totals); three renderers turn that into bytes.
Keeping the table model separate from the renderers means a new report only has to
describe its columns once, and a new format only has to be written once.

Money is rendered as human-facing naira strings in exports (the JSON API carries the
exact kobo); callers pass already-formatted strings in the cells.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field

#: Supported export formats → (extension, MIME content type).
EXPORT_FORMATS = {  # Public map of supported export format metadata.
    "csv": ("csv", "text/csv"),  # CSV extension and MIME type.
    "xlsx": ("xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),  # Excel extension and MIME type.
    "pdf": ("pdf", "application/pdf"),  # PDF extension and MIME type.
}


@dataclass
# Renderer-neutral table representation.
class ReportTable:
    """A renderer-neutral rectangular report.

    ``rows`` and ``summary_rows`` are lists of cells (str/number); ``summary_rows`` are
    rendered emphasised (bold) and separated, for totals lines. ``subtitle`` is an
    optional second heading line (e.g. the entity + period).
    """

    title: str  # Main report heading.
    columns: list[str]  # Column labels shared by all renderers.
    rows: list[list] = field(default_factory=list)  # Body rows.
    summary_rows: list[list] = field(default_factory=list)  # Emphasized total/summary rows.
    subtitle: str = ""  # Optional second heading line.

    @property
    # Combine body and summary rows for renderers.
    def all_rows(self) -> list[list]:
        return list(self.rows) + list(self.summary_rows)  # Return copies to avoid mutating source lists.


# Render a report table to UTF-8 CSV bytes.
def to_csv(table: ReportTable) -> bytes:
    buf = io.StringIO()  # Hold CSV text before encoding.
    writer = csv.writer(buf)  # Use Python's CSV escaping rules.
    writer.writerow([table.title])  # First row is the report title.
    if table.subtitle:  # Include subtitle only when provided.
        writer.writerow([table.subtitle])  # Second row is the subtitle.
    writer.writerow([])  # Blank separator before headers.
    writer.writerow(table.columns)  # Write column headers.
    for row in table.rows:  # Write normal body rows.
        writer.writerow(row)  # Preserve caller-supplied cell values.
    if table.summary_rows:  # Add a visual break before totals.
        writer.writerow([])  # Blank separator before summary rows.
        for row in table.summary_rows:  # Write emphasized rows without CSV styling.
            writer.writerow(row)  # CSV cannot carry bold style, so values only.
    return buf.getvalue().encode("utf-8")  # Return encoded bytes for HTTP response.


# Render a report table to Excel workbook bytes.
def to_xlsx(table: ReportTable) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()  # Create a single-workbook export.
    ws = wb.active  # Use the default worksheet.
    ws.title = (table.title[:28] or "Report")  # Excel sheet names are limited; keep a safe title.

    bold = Font(bold=True)  # Reusable bold font style.
    ws.append([table.title])  # First row is the report title.
    ws["A1"].font = bold  # Emphasize the title.
    if table.subtitle:  # Include subtitle only when provided.
        ws.append([table.subtitle])  # Second row is the subtitle.
    ws.append([])  # Blank separator before headers.

    header_row = ws.max_row + 1  # Capture the row index that will hold headers.
    ws.append(table.columns)  # Write column labels.
    for cell in ws[header_row]:  # Style every header cell.
        cell.font = bold  # Make headers bold.

    for row in table.rows:  # Append body rows.
        ws.append(row)  # Preserve caller-supplied values.
    for row in table.summary_rows:  # Append total/summary rows.
        ws.append(row)  # Add summary row values.
        for cell in ws[ws.max_row]:  # Style the row just appended.
            cell.font = bold  # Emphasize summary rows.

    # Roughly autosize columns to their widest cell.  # Improves readability without complex layout.
    for idx, _col in enumerate(table.columns, start=1):  # Walk each report column.
        from openpyxl.utils import get_column_letter
        letter = get_column_letter(idx)  # Convert 1-based index to Excel column letter.
        widest = max(  # Compute the widest visible value in this column.
            [len(str(table.columns[idx - 1]))]  # Include header width.
            + [len(str(r[idx - 1])) for r in table.all_rows if idx - 1 < len(r)]  # Include row cell widths when present.
            or [10]  # Defensive fallback width.
        )
        ws.column_dimensions[letter].width = min(max(widest + 2, 10), 48)  # Clamp to a usable width range.

    out = io.BytesIO()  # Store workbook bytes in memory.
    wb.save(out)
    return out.getvalue()  # Return xlsx bytes for HTTP response.


# Render a report table to PDF bytes.
def to_pdf(table: ReportTable) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    out = io.BytesIO()  # Store PDF bytes in memory.
    doc = SimpleDocTemplate(out, pagesize=landscape(A4), title=table.title)  # Configure landscape PDF document.
    styles = getSampleStyleSheet()  # Load ReportLab default text styles.
    story = [Paragraph(table.title, styles["Title"])]  # Start document with title.
    if table.subtitle:  # Include subtitle only when provided.
        story.append(Paragraph(table.subtitle, styles["Normal"]))  # Add subtitle under the title.
    story.append(Spacer(1, 12))  # Add spacing before the table.

    data = [table.columns] + [[str(c) for c in r] for r in table.all_rows]  # Convert all cells to PDF-safe strings.
    pdf_table = Table(data, repeatRows=1)  # Repeat headers across pages.
    style = [  # Base table style.
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),  # Dark header background.
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),  # White header text.
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),  # Bold header font.
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),  # Light grid around cells.
        ("FONTSIZE", (0, 0), (-1, -1), 8),  # Compact report font size.
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f4f6")]),  # Alternating body backgrounds.
    ]
    # Bold the summary rows at the bottom.  # Makes totals stand out.
    n_summary = len(table.summary_rows)  # Count summary rows for bottom styling.
    if n_summary:  # Apply summary styling only when summary rows exist.
        first = len(data) - n_summary  # First summary row index in the PDF table.
        style.append(("FONTNAME", (0, first), (-1, -1), "Helvetica-Bold"))  # Bold every summary cell.
        style.append(("LINEABOVE", (0, first), (-1, first), 0.8, colors.black))  # Draw separator above totals.
    pdf_table.setStyle(TableStyle(style))  # Attach style instructions to the table.
    story.append(pdf_table)  # Add the table to the PDF story.

    doc.build(story)  # Render the PDF document into the buffer.
    return out.getvalue()  # Return PDF bytes for HTTP response.


_RENDERERS = {"csv": to_csv, "xlsx": to_xlsx, "pdf": to_pdf}  # Dispatch table for supported renderers.


# Render a report in the requested format.
def render(table: ReportTable, fmt: str) -> tuple[bytes, str, str]:
    """Render ``table`` in ``fmt`` → ``(body, content_type, file_extension)``.

    Raises :class:`ValueError` for an unsupported format (the view turns this into a
    400). The filename is the caller's concern; only the extension is returned here.
    """
    fmt = (fmt or "").lower()  # Normalize missing and mixed-case format values.
    if fmt not in _RENDERERS:  # Reject unsupported export formats early.
        raise ValueError(
            f"Unsupported export format '{fmt}'. Choose one of: {', '.join(EXPORT_FORMATS)}."
        )
    extension, content_type = EXPORT_FORMATS[fmt]  # Look up HTTP metadata for this format.
    return _RENDERERS[fmt](table), content_type, extension  # Render and return body plus metadata.
