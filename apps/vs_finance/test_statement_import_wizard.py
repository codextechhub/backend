from __future__ import annotations

from io import BytesIO
from unittest import mock

from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.test import SimpleTestCase, TestCase

from core.test_utils import TenantAPIClient
from vs_import_data.models import (
    DatasetTypeChoices,
    FileFormatChoices,
    ImportBatch,
    ImportBatchStatusChoices,
    ImportJob,
    ImportJobStatusChoices,
    ImportTemplate,
    ImportTemplateColumn,
    TemplateColumnDataTypeChoices,
)
from vs_rbac.tests.helpers import (
    codex_tenant,
    make_assignment,
    make_permission,
    make_role,
    make_role_permission,
    make_vision_user,
)

from .constants import BankLineStatus, FinanceAuditAction
from .models import (
    BankStatement,
    BankStatementImportContext,
    BankStatementLine,
    FinanceAuditLog,
)
from .tests import _Phase4FixtureMixin


HEADERS = (
    "Transaction Date,Description,Reference,Money In,Money Out,"
    "Transaction ID,Balance\n"
)


class BankStatementImportWizardTests(_Phase4FixtureMixin, TestCase):
    def setUp(self):
        self.user = make_vision_user(
            email="statement-import@test.com",
            super_admin=True,
        )
        self.client = TenantAPIClient(user=self.user)
        call_command(
            "seed_import",
            dataset_type=DatasetTypeChoices.BANK_STATEMENTS,
            verbosity=0,
        )
        self.entity, _, _ = self.build_books()
        self.bank = self.make_bank(self.entity)

    def _file(self, body: str, name="statement.csv"):
        return SimpleUploadedFile(
            name,
            (HEADERS + body).encode(),
            content_type="text/csv",
        )

    def _upload(
        self,
        body: str,
        *,
        opening="1000.00",
        closing="1450.00",
        name="statement.csv",
        client=None,
        bank=None,
        entity=None,
    ):
        client = client or self.client
        bank = bank or self.bank
        entity = entity or self.entity
        return client.post(
            f"/v1/finance/bank-accounts/{bank.pk}/statement-imports/?entity={entity.code}",
            {
                "file": self._file(body, name=name),
                "statement_date": "2026-01-31",
                "period_label": "January 2026",
                "opening_balance": opening,
                "closing_balance": closing,
            },
            format="multipart",
        )

    def _balanced_rows(self):
        return (
            "2026-01-02,Customer receipt,RCPT-1,500.00,,TX-1,1500.00\n"
            "2026-01-03,Bank fee,FEE-1,,50.00,TX-2,1450.00\n"
        )

    def _manual_statement(self):
        response = self.client.post(
            (
                f"/v1/finance/bank-accounts/{self.bank.pk}/statement-lines/"
                f"?entity={self.entity.code}"
            ),
            {
                "statement_date": "2026-01-31",
                "period_label": "January 2026",
                "opening_balance": 100000,
                "lines": [
                    {
                        "txn_date": "2026-01-02",
                        "description": "Customer receipt",
                        "reference": "RCPT-1",
                        "amount": 50000,
                    },
                    {
                        "txn_date": "2026-01-03",
                        "description": "Bank fee",
                        "reference": "FEE-1",
                        "amount": -5000,
                    },
                ],
            },
            format="json",
        )
        self.assertEqual(response.status_code, 201, response.content)
        return BankStatement.objects.get(
            pk=response.json()["data"]["imported"][0]["statement_id"],
        )

    def test_upload_validates_then_publish_creates_one_statement_atomically(self):
        upload = self._upload(self._balanced_rows())
        self.assertEqual(upload.status_code, 201, upload.content)
        data = upload.json()["data"]
        self.assertEqual(data["status"], ImportBatchStatusChoices.READY_TO_IMPORT)
        self.assertTrue(data["is_ready_for_import"])
        self.assertEqual(data["validation_summary"]["error_count"], 0)
        self.assertEqual(data["validation_summary"]["dataset"]["money_in"], 50000)
        self.assertEqual(data["validation_summary"]["dataset"]["money_out"], 5000)
        self.assertEqual(data["domain_context"]["bank_account_id"], self.bank.pk)
        batch_id = data["id"]

        publish = self.client.post(
            f"/v1/import/batches/{batch_id}/start-import/",
            {"run_async": False},
            format="json",
        )
        self.assertEqual(publish.status_code, 200, publish.content)

        context = BankStatementImportContext.objects.get(import_batch_id=batch_id)
        self.assertIsNotNone(context.published_statement_id)
        statement = BankStatement.objects.get(pk=context.published_statement_id)
        self.assertEqual(statement.opening_balance, 100000)
        self.assertEqual(statement.closing_balance, 145000)
        self.assertEqual(
            list(statement.lines.order_by("txn_date").values_list("amount", flat=True)),
            [50000, -5000],
        )

    def test_publish_finalization_failure_rolls_back_the_statement(self):
        upload = self._upload(self._balanced_rows())
        batch_id = upload.json()["data"]["id"]

        with mock.patch(
            "vs_import_data.services.import_executor.finalize_import_job",
            side_effect=RuntimeError("Audit store unavailable"),
        ):
            publish = self.client.post(
                f"/v1/import/batches/{batch_id}/start-import/",
                {"run_async": False},
                format="json",
            )

        self.assertEqual(publish.status_code, 400, publish.content)
        self.assertFalse(BankStatement.objects.exists())
        self.assertFalse(BankStatementLine.objects.exists())
        self.assertIsNone(
            BankStatementImportContext.objects.get(
                import_batch_id=batch_id,
            ).published_statement_id
        )
        self.assertEqual(
            ImportJob.objects.get(import_batch_id=batch_id).status,
            ImportJobStatusChoices.FAILED,
        )

    def test_imbalanced_statement_is_blocked_before_publish(self):
        upload = self._upload(self._balanced_rows(), closing="1400.00")
        self.assertEqual(upload.status_code, 201, upload.content)
        data = upload.json()["data"]
        self.assertEqual(data["status"], ImportBatchStatusChoices.VALIDATION_FAILED)
        self.assertFalse(data["is_ready_for_import"])
        self.assertTrue(any(
            "does not equal" in issue["message"]
            for issue in data["validation_issues"]
        ))

        publish = self.client.post(
            f"/v1/import/batches/{data['id']}/start-import/",
            {"run_async": False},
            format="json",
        )
        self.assertEqual(publish.status_code, 400, publish.content)
        self.assertFalse(BankStatement.objects.exists())
        self.assertFalse(BankStatementLine.objects.exists())

    def test_duplicate_file_and_transactions_are_blocked_on_second_wizard(self):
        first = self._upload(self._balanced_rows())
        first_id = first.json()["data"]["id"]
        published = self.client.post(
            f"/v1/import/batches/{first_id}/start-import/",
            {"run_async": False},
            format="json",
        )
        self.assertEqual(published.status_code, 200, published.content)

        second = self._upload(self._balanced_rows())
        self.assertEqual(second.status_code, 201, second.content)
        data = second.json()["data"]
        self.assertFalse(data["is_ready_for_import"])
        messages = [issue["message"] for issue in data["validation_issues"]]
        self.assertTrue(any("exact statement file" in message for message in messages))
        self.assertTrue(any("already been imported" in message for message in messages))
        self.assertEqual(BankStatement.objects.count(), 1)
        self.assertEqual(BankStatementLine.objects.count(), 2)

    def test_running_balance_and_date_order_are_validated(self):
        rows = (
            "2026-01-03,Customer receipt,RCPT-1,500.00,,TX-1,1499.00\n"
            "2026-01-02,Bank fee,FEE-1,,50.00,TX-2,1450.00\n"
        )
        upload = self._upload(rows)
        data = upload.json()["data"]
        self.assertFalse(data["is_ready_for_import"])
        messages = [issue["message"] for issue in data["validation_issues"]]
        self.assertTrue(any("oldest to newest" in message for message in messages))
        self.assertTrue(any("Running balance" in message for message in messages))

    def test_template_download_uses_the_same_wizard_format(self):
        template_url = (
            f"/v1/finance/bank-accounts/{self.bank.pk}/statement-imports/template/"
            f"?entity={self.entity.code}"
        )
        csv_response = self.client.get(f"{template_url}&file_format=csv")
        self.assertEqual(csv_response.status_code, 200, csv_response.content)
        self.assertEqual(csv_response["Content-Type"], "text/csv")
        csv_content = (
            b"".join(csv_response.streaming_content)
            if csv_response.streaming
            else csv_response.content
        )
        csv_rows = csv_content.decode().splitlines()
        self.assertEqual(csv_rows[0], HEADERS.strip())
        self.assertEqual(
            csv_rows[1],
            "2026-07-01,Customer payment,BANK-REF-0001,1250.00,,"
            "TXN-000001,11250.00",
        )

        xlsx_response = self.client.get(f"{template_url}&file_format=xlsx")
        self.assertEqual(xlsx_response.status_code, 200, xlsx_response.content)
        self.assertEqual(
            xlsx_response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        xlsx_content = (
            b"".join(xlsx_response.streaming_content)
            if xlsx_response.streaming
            else xlsx_response.content
        )

        from openpyxl import load_workbook

        sheet = load_workbook(BytesIO(xlsx_content), read_only=True).worksheets[0]
        rows = list(sheet.iter_rows(max_row=2, values_only=True))
        self.assertEqual(rows[0], tuple(HEADERS.strip().split(",")))
        self.assertEqual(
            rows[1],
            (
                "2026-07-01",
                "Customer payment",
                "BANK-REF-0001",
                "1250.00",
                None,
                "TXN-000001",
                "11250.00",
            ),
        )

    def test_excel_date_cells_are_normalized_by_the_wizard(self):
        from datetime import datetime
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(HEADERS.strip().split(","))
        sheet.append([
            datetime(2026, 1, 2),
            "Customer receipt",
            "RCPT-1",
            500,
            None,
            "TX-1",
            1500,
        ])
        sheet.append([
            datetime(2026, 1, 3),
            "Bank fee",
            "FEE-1",
            None,
            50,
            "TX-2",
            1450,
        ])
        contents = BytesIO()
        workbook.save(contents)
        workbook.close()
        contents.seek(0)

        response = self.client.post(
            (
                f"/v1/finance/bank-accounts/{self.bank.pk}/statement-imports/"
                f"?entity={self.entity.code}"
            ),
            {
                "file": SimpleUploadedFile(
                    "statement.xlsx",
                    contents.read(),
                    content_type=(
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet"
                    ),
                ),
                "statement_date": "2026-01-31",
                "opening_balance": "1000.00",
                "closing_balance": "1450.00",
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, 201, response.content)
        data = response.json()["data"]
        self.assertTrue(data["is_ready_for_import"], data["validation_issues"])
        self.assertEqual(data["preview_rows"][0]["Transaction Date"], "2026-01-02")

    def test_unreconciled_published_statement_can_be_rolled_back(self):
        upload = self._upload(self._balanced_rows())
        batch_id = upload.json()["data"]["id"]
        publish = self.client.post(
            f"/v1/import/batches/{batch_id}/start-import/",
            {"run_async": False},
            format="json",
        )
        job_id = publish.json()["data"]["job_id"]

        rolled_back = self.client.post(
            f"/v1/import/batches/{batch_id}/jobs/{job_id}/rollback/",
            {"reason": "Wrong statement period"},
            format="json",
        )
        self.assertEqual(rolled_back.status_code, 200, rolled_back.content)
        self.assertEqual(rolled_back.json()["data"]["reverted_rows_count"], 2)
        self.assertFalse(BankStatement.objects.exists())
        self.assertFalse(BankStatementLine.objects.exists())
        context = BankStatementImportContext.objects.get(import_batch_id=batch_id)
        self.assertIsNone(context.published_statement_id)

    def test_statement_with_an_acted_on_line_cannot_be_rolled_back(self):
        upload = self._upload(self._balanced_rows())
        batch_id = upload.json()["data"]["id"]
        publish = self.client.post(
            f"/v1/import/batches/{batch_id}/start-import/",
            {"run_async": False},
            format="json",
        )
        job_id = publish.json()["data"]["job_id"]
        delete_published = self.client.delete(f"/v1/import/batches/{batch_id}/")
        self.assertEqual(delete_published.status_code, 400, delete_published.content)
        line = BankStatementLine.objects.order_by("id").first()
        ignored = self.client.post(
            f"/v1/finance/statement-lines/{line.pk}/ignore/?entity={self.entity.code}",
            {"ignored": True, "reason": "Known bank-only line"},
            format="json",
        )
        self.assertEqual(ignored.status_code, 200, ignored.content)

        rollback = self.client.post(
            f"/v1/import/batches/{batch_id}/jobs/{job_id}/rollback/",
            {"reason": "Try after action"},
            format="json",
        )
        self.assertEqual(rollback.status_code, 400, rollback.content)
        self.assertTrue(BankStatement.objects.exists())
        self.assertEqual(BankStatementLine.objects.count(), 2)

    def test_manual_statement_can_be_corrected_before_reconciliation(self):
        statement = self._manual_statement()
        detail_url = (
            f"/v1/finance/bank-accounts/{self.bank.pk}/statements/{statement.pk}/"
            f"?entity={self.entity.code}"
        )
        detail = self.client.get(detail_url)
        self.assertEqual(detail.status_code, 200, detail.content)
        data = detail.json()["data"]
        self.assertTrue(data["can_edit"])
        first_line, second_line = data["lines"]

        corrected = self.client.patch(
            detail_url,
            {
                "statement_date": "2026-02-01",
                "period_label": "Corrected January 2026",
                "opening_balance": 110000,
                "lines": [
                    {
                        "id": first_line["id"],
                        "txn_date": "2026-01-02",
                        "description": "Corrected customer receipt",
                        "reference": "RCPT-1",
                        "amount": 60000,
                        "external_id": "",
                    },
                    {
                        "txn_date": "2026-01-04",
                        "description": "Replacement bank fee",
                        "reference": "FEE-2",
                        "amount": -10000,
                        "external_id": "",
                    },
                ],
            },
            format="json",
        )
        self.assertEqual(corrected.status_code, 200, corrected.content)
        corrected_data = corrected.json()["data"]
        self.assertEqual(corrected_data["closing_balance"], 160000)
        self.assertEqual(corrected_data["line_count"], 2)
        self.assertFalse(BankStatementLine.objects.filter(pk=second_line["id"]).exists())
        self.assertEqual(
            list(
                statement.lines.order_by("txn_date").values_list(
                    "description",
                    "amount",
                )
            ),
            [
                ("Corrected customer receipt", 60000),
                ("Replacement bank fee", -10000),
            ],
        )
        audit = FinanceAuditLog.objects.get(
            action=FinanceAuditAction.BANK_STATEMENT_CORRECTED,
            target_id=str(statement.pk),
        )
        self.assertEqual(audit.actor_id, self.user.pk)
        self.assertEqual(audit.before["closing_balance"], 145000)
        self.assertEqual(audit.after["closing_balance"], 160000)

    def test_manual_statement_correction_is_blocked_after_a_line_is_acted_on(self):
        statement = self._manual_statement()
        line = statement.lines.first()
        line.status = BankLineStatus.IGNORED
        line.save(update_fields=["status", "updated_at"])

        response = self.client.patch(
            (
                f"/v1/finance/bank-accounts/{self.bank.pk}/statements/{statement.pk}/"
                f"?entity={self.entity.code}"
            ),
            {
                "statement_date": "2026-01-31",
                "opening_balance": 100000,
                "lines": [
                    {
                        "id": existing.pk,
                        "txn_date": existing.txn_date,
                        "description": existing.description,
                        "reference": existing.reference,
                        "amount": existing.amount,
                    }
                    for existing in statement.lines.all()
                ],
            },
            format="json",
        )
        self.assertEqual(response.status_code, 400, response.content)
        self.assertIn("restore every statement line", str(response.json()).lower())

    def test_unmatched_manual_lines_can_be_deleted_and_empty_statement_is_removed(self):
        statement = self._manual_statement()
        receipt, fee = list(statement.lines.order_by("txn_date", "id"))

        deleted = self.client.delete(
            (
                f"/v1/finance/statement-lines/{receipt.pk}/"
                f"?entity={self.entity.code}"
            ),
        )
        self.assertEqual(deleted.status_code, 200, deleted.content)
        self.assertIsNone(deleted.json()["data"]["deleted_statement_id"])
        statement.refresh_from_db()
        self.assertEqual(statement.closing_balance, 95000)
        self.assertEqual(list(statement.lines.values_list("pk", flat=True)), [fee.pk])

        deleted_final = self.client.delete(
            (
                f"/v1/finance/statement-lines/{fee.pk}/"
                f"?entity={self.entity.code}"
            ),
        )
        self.assertEqual(deleted_final.status_code, 200, deleted_final.content)
        self.assertEqual(
            deleted_final.json()["data"]["deleted_statement_id"],
            statement.pk,
        )
        self.assertFalse(BankStatement.objects.filter(pk=statement.pk).exists())
        self.assertFalse(BankStatementLine.objects.filter(pk=fee.pk).exists())
        self.assertEqual(
            FinanceAuditLog.objects.filter(
                action=FinanceAuditAction.BANK_STATEMENT_CORRECTED,
                metadata__deleted_line_id__in=[receipt.pk, fee.pk],
            ).count(),
            2,
        )

    def test_acted_on_statement_line_cannot_be_deleted(self):
        statement = self._manual_statement()
        line = statement.lines.first()
        line.status = BankLineStatus.IGNORED
        line.save(update_fields=["status", "updated_at"])

        response = self.client.delete(
            (
                f"/v1/finance/statement-lines/{line.pk}/"
                f"?entity={self.entity.code}"
            ),
        )
        self.assertEqual(response.status_code, 400, response.content)
        self.assertTrue(BankStatementLine.objects.filter(pk=line.pk).exists())

    def test_bulk_statement_correction_requires_rollback_and_reimport(self):
        upload = self._upload(self._balanced_rows())
        batch_id = upload.json()["data"]["id"]
        publish = self.client.post(
            f"/v1/import/batches/{batch_id}/start-import/",
            {"run_async": False},
            format="json",
        )
        self.assertEqual(publish.status_code, 200, publish.content)
        statement = BankStatement.objects.get(
            pk=BankStatementImportContext.objects.get(
                import_batch_id=batch_id,
            ).published_statement_id,
        )
        detail_url = (
            f"/v1/finance/bank-accounts/{self.bank.pk}/statements/{statement.pk}/"
            f"?entity={self.entity.code}"
        )
        detail = self.client.get(detail_url)
        self.assertFalse(detail.json()["data"]["can_edit"])
        self.assertIn(
            "rolled back and re-imported",
            detail.json()["data"]["edit_block_reason"],
        )

        response = self.client.patch(
            detail_url,
            {
                "statement_date": "2026-01-31",
                "opening_balance": 100000,
                "lines": [
                    {
                        "id": line.pk,
                        "txn_date": line.txn_date,
                        "description": line.description,
                        "reference": line.reference,
                        "amount": line.amount,
                        "external_id": line.external_id,
                    }
                    for line in statement.lines.all()
                ],
            },
            format="json",
        )
        self.assertEqual(response.status_code, 400, response.content)
        self.assertIn("rolled back and re-imported", str(response.json()))

    def test_user_without_import_permission_cannot_correct_manual_statement(self):
        statement = self._manual_statement()
        user = make_vision_user(email="no-statement-correction@test.com")
        client = TenantAPIClient(user=user)
        response = client.patch(
            (
                f"/v1/finance/bank-accounts/{self.bank.pk}/statements/{statement.pk}/"
                f"?entity={self.entity.code}"
            ),
            {},
            format="json",
        )
        self.assertEqual(response.status_code, 403, response.content)

    def test_finance_import_permission_opens_only_its_bank_statement_wizard(self):
        tenant = codex_tenant()
        finance_user = make_vision_user(
            email="finance-import-only@test.com",
            tenant=tenant,
        )
        role = make_role(tenant, name="Statement Importer")
        make_role_permission(role, make_permission("finance.bankaccount.import"))
        make_assignment(tenant, finance_user, role)
        client = TenantAPIClient(user=finance_user)

        upload = self._upload(self._balanced_rows(), client=client)
        self.assertEqual(upload.status_code, 201, upload.content)
        batch_id = upload.json()["data"]["id"]
        detail = client.get(f"/v1/import/batches/{batch_id}/")
        self.assertEqual(detail.status_code, 200, detail.content)
        self.assertIn("preview_rows", detail.json()["data"])
        publish = client.post(
            f"/v1/import/batches/{batch_id}/start-import/",
            {"run_async": False},
            format="json",
        )
        self.assertEqual(publish.status_code, 200, publish.content)

        other_template = ImportTemplate.objects.create(
            code="permission-boundary-test",
            name="Permission Boundary",
            dataset_type=DatasetTypeChoices.SCHOOLS,
            default_file_format=FileFormatChoices.CSV,
        )
        ImportTemplateColumn.objects.create(
            template=other_template,
            column_name="Name",
            target_field="name",
            data_type=TemplateColumnDataTypeChoices.STRING,
            is_required=True,
        )
        other_batch = ImportBatch.objects.create(
            tenant=tenant,
            uploaded_by=self.user,
            template=other_template,
            dataset_type=DatasetTypeChoices.SCHOOLS,
            file=SimpleUploadedFile("other.csv", b"Name\nOther\n"),
            file_format=FileFormatChoices.CSV,
            original_filename="other.csv",
            total_rows=1,
            total_columns=1,
            uploaded_headers=["Name"],
            preview_rows=[{"Name": "Other"}],
        )
        denied = client.get(f"/v1/import/batches/{other_batch.pk}/")
        self.assertEqual(denied.status_code, 403, denied.content)

    def test_user_without_finance_import_permission_cannot_upload(self):
        user = make_vision_user(email="no-statement-import@test.com")
        client = TenantAPIClient(user=user)
        response = self._upload(self._balanced_rows(), client=client)
        self.assertEqual(response.status_code, 403, response.content)

    def test_finance_importer_can_cancel_own_statement_batch(self):
        tenant = codex_tenant()
        user = make_vision_user(email="statement-cancel@test.com", tenant=tenant)
        role = make_role(tenant, name="Statement cancellation importer")
        make_role_permission(role, make_permission("finance.bankaccount.import"))
        make_assignment(tenant, user, role)
        client = TenantAPIClient(user=user)
        upload = self._upload(self._balanced_rows(), client=client)
        batch_id = upload.json()["data"]["id"]

        cancelled = client.post(f"/v1/import/batches/{batch_id}/cancel/")

        self.assertEqual(cancelled.status_code, 200, cancelled.content)
        self.assertEqual(
            ImportBatch.objects.get(pk=batch_id).status,
            ImportBatchStatusChoices.CANCELLED,
        )

    def test_cross_tenant_bank_account_is_not_uploadable(self):
        from vs_tenants.models import Tenant
        from vs_finance.models import Account, BankAccount, LedgerEntity
        from vs_finance.seed import seed_chart_of_accounts

        other = Tenant.objects.create(name="Other Tenant", slug="other-statement")
        foreign_entity = LedgerEntity.objects.create(
            tenant=other,
            name="Other Books",
            code="OTHERBANK",
            kind=LedgerEntity.Kind.TENANT,
        )
        seed_chart_of_accounts(foreign_entity)
        foreign_bank = BankAccount.objects.create(
            entity=foreign_entity,
            name="Foreign Bank",
            gl_account=Account.objects.get(entity=foreign_entity, code="1100"),
        )

        response = self.client.post(
            (
                f"/v1/finance/bank-accounts/{foreign_bank.pk}/statement-imports/"
                f"?entity={foreign_entity.code}"
            ),
            {
                "file": self._file(self._balanced_rows()),
                "statement_date": "2026-01-31",
                "opening_balance": "1000.00",
                "closing_balance": "1450.00",
            },
            format="multipart",
        )
        self.assertEqual(response.status_code, 404, response.content)

        foreign_statement = BankStatement.objects.create(
            bank_account=foreign_bank,
            statement_date="2026-01-31",
            opening_balance=0,
            closing_balance=10000,
        )
        foreign_line = BankStatementLine.objects.create(
            bank_account=foreign_bank,
            statement=foreign_statement,
            txn_date="2026-01-02",
            amount=10000,
        )
        correction = self.client.patch(
            (
                f"/v1/finance/bank-accounts/{foreign_bank.pk}/statements/"
                f"{foreign_statement.pk}/?entity={foreign_entity.code}"
            ),
            {
                "statement_date": "2026-01-31",
                "opening_balance": 0,
                "lines": [
                    {
                        "id": foreign_line.pk,
                        "txn_date": "2026-01-02",
                        "amount": 20000,
                    },
                ],
            },
            format="json",
        )
        self.assertEqual(correction.status_code, 404, correction.content)


class ImportFileParserSafetyTests(SimpleTestCase):
    def test_row_limit_is_explicit_instead_of_silent_truncation(self):
        from vs_import_data.services.file_parser import parse_csv

        raw = BytesIO(b"Name\nOne\nTwo\nThree\n")
        with mock.patch("vs_import_data.services.file_parser.MAX_IMPORT_ROWS", 2):
            with self.assertRaisesRegex(ValueError, "more than 2 data rows"):
                parse_csv(raw)
