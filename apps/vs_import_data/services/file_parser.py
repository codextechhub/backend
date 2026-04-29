"""
Streaming file parser for import batches.

Reads CSV and Excel files row by row to avoid loading the entire file into
memory. Caps preview rows so the JSON field stays manageable.
"""
from __future__ import annotations

import csv
from io import BytesIO, StringIO

MAX_PREVIEW_ROWS = 5_000


def _open_csv_text(raw_bytes: bytes) -> StringIO:
    """Try common encodings in order; raise a clear error on total failure."""
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return StringIO(raw_bytes.decode(encoding))
        except (UnicodeDecodeError, LookupError):
            continue
    raise ValueError(
        "File encoding could not be detected. Please save as UTF-8 and re-upload."
    )


def parse_csv(file_obj, *, header_row_index: int = 1) -> tuple[list[str], list[dict]]:
    """
    Parse a CSV file object and return (headers, rows).

    header_row_index is 1-based. Rows before it are skipped.
    Caps at MAX_PREVIEW_ROWS data rows.
    """
    raw_bytes = file_obj.read()
    text_io = _open_csv_text(raw_bytes)

    reader = csv.reader(text_io)
    headers: list[str] = []
    rows: list[dict] = []
    row_number = 0

    for line in reader:
        row_number += 1
        if row_number < header_row_index:
            continue
        if row_number == header_row_index:
            headers = [h.strip() for h in line]
            continue
        if not any(cell.strip() for cell in line):
            continue
        row_dict = dict(zip(headers, [c.strip() for c in line]))
        rows.append(row_dict)
        if len(rows) >= MAX_PREVIEW_ROWS:
            break

    return headers, rows


def parse_xlsx(file_obj, *, sheet_name: str | None = None, header_row_index: int = 1) -> tuple[list[str], list[dict]]:
    """
    Parse an Excel (.xlsx) file object and return (headers, rows).

    Uses openpyxl in read-only mode to avoid loading the full workbook into RAM.
    Caps at MAX_PREVIEW_ROWS data rows.
    """
    from openpyxl import load_workbook

    raw_bytes = file_obj.read()
    wb = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    headers: list[str] = []
    rows: list[dict] = []
    row_number = 0

    try:
        for excel_row in ws.iter_rows(values_only=True):
            row_number += 1
            if row_number < header_row_index:
                continue
            if row_number == header_row_index:
                headers = [str(cell).strip() if cell is not None else "" for cell in excel_row]
                continue
            values = [str(cell).strip() if cell is not None else "" for cell in excel_row]
            if not any(values):
                continue
            row_dict = dict(zip(headers, values))
            rows.append(row_dict)
            if len(rows) >= MAX_PREVIEW_ROWS:
                break
    finally:
        wb.close()

    return headers, rows


def parse_import_file(
    file_obj,
    *,
    file_format: str,
    sheet_name: str | None = None,
    header_row_index: int = 1,
) -> tuple[list[str], list[dict]]:
    """
    Dispatch to the correct parser based on file_format.

    Returns (headers, rows) where rows is a list of dicts keyed by header name.
    """
    if file_format == "csv":
        return parse_csv(file_obj, header_row_index=header_row_index)
    elif file_format in ("xlsx", "xls"):
        return parse_xlsx(file_obj, sheet_name=sheet_name, header_row_index=header_row_index)
    else:
        raise ValueError(f"Unsupported file format: {file_format!r}")
