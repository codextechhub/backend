from __future__ import annotations

import csv
from io import BytesIO, StringIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from ..models import FileFormatChoices, ImportBatch, ImportTemplate


def generate_template_xlsx(template: ImportTemplate) -> bytes:
    """
    Generate an Excel template file from ImportTemplate and ImportTemplateColumn.
    """
    wb = Workbook()

    # Main data sheet
    ws = wb.active
    ws.title = template.dataset_type[:31]  # Excel sheet name limit

    columns = list(template.columns.order_by("column_order"))

    headers = [col.column_name for col in columns]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    bold = Font(bold=True)

    for cell in ws[1]:
        cell.font = bold
        cell.fill = header_fill

    if template.allow_sample_row:
        sample_row = []
        for col in columns:
            value = (
                template.sample_row_data.get(col.column_name)
                or col.sample_value
                or col.default_value
                or ""
            )
            sample_row.append(value)
        ws.append(sample_row)

    # Instructions sheet
    info = wb.create_sheet(title="Instructions")
    info.append(["Template Name", template.name])
    info.append(["Dataset Type", template.dataset_type])
    info.append(["Version", template.version])
    info.append(["Description", template.description or ""])
    info.append(["Instructions", template.instructions or ""])

    info.append([])
    info.append([
        "Column Name",
        "Target Field",
        "Required",
        "Data Type",
        "Allowed Values",
        "Help Text",
        "Sample Value",
    ])

    for col in columns:
        info.append([
            col.column_name,
            col.target_field,
            "Yes" if col.is_required else "No",
            col.data_type,
            ", ".join(col.allowed_values or []),
            col.help_text,
            col.sample_value,
        ])

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_template_csv(template: ImportTemplate) -> str:
    """
    Generate a CSV template file from ImportTemplate and ImportTemplateColumn.
    """
    output = StringIO()
    writer = csv.writer(output)

    columns = list(template.columns.order_by("column_order"))
    headers = [col.column_name for col in columns]
    writer.writerow(headers)

    if template.allow_sample_row:
        sample_row = []
        for col in columns:
            value = (
                template.sample_row_data.get(col.column_name)
                or col.sample_value
                or col.default_value
                or ""
            )
            sample_row.append(value)
        writer.writerow(sample_row)

    return output.getvalue()


def generate_validation_issues_csv(import_batch: ImportBatch) -> str:
    """
    Generate a CSV of all validation issues for an import batch.
    File-level issues (no row) appear first, then row issues sorted by row number.
    """
    output = StringIO()
    writer = csv.writer(output)

    writer.writerow([
        "Row",
        "Column",
        "Severity",
        "Error Code",
        "Message",
        "Raw Value",
        "Help Text",
        "Resolved",
    ])

    issues = import_batch.validation_issues.order_by("row_number", "column_name", "created_at")

    for issue in issues:
        writer.writerow([
            issue.row_number if issue.row_number is not None else "File",
            issue.column_name or "",
            issue.severity.upper(),
            issue.code,
            issue.message,
            issue.raw_value or "",
            issue.help_text or "",
            "Yes" if issue.is_resolved else "No",
        ])

    return output.getvalue()